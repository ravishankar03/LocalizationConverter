
from openpyxl import load_workbook
from openpyxl import Workbook
from googletrans import Translator
import os

def translate_excel(file_path, columns_to_translate, source_lang="en", target_lang="es", sheet_name=None, replace=False):
    """
    Translate specific columns in an Excel file and save the output with a language suffix.

    Parameters:
    - file_path: Path to the Excel file.
    - columns_to_translate: List of column numbers (1-based) to translate.
    - source_lang: Source language code (default: 'en').
    - target_lang: Target language code (default: 'es').
    - sheet_name: Specific sheet name (optional, default: active sheet).
    - replace: Whether to replace the existing columns (default: False).
    """
    # Load Excel file
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    # Initialize translator
    translator = Translator()

    for source_col in columns_to_translate:
        # Determine where to write translations
        if replace:
            print(f"Replacing column {source_col} with translations...")
            target_col = source_col  # Replace source column
        else:
            target_col = sheet.max_column + 1  # Add translations to a new column
            print(f"Adding translations from column {source_col} to a new column ({target_col})...")

        # Iterate through rows and translate
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # Assuming row 1 is header
            source_text = row[source_col - 1].value  # Adjust for 0-based index
            if source_text:
                try:
                    # Translate text
                    translation = translator.translate(source_text, src=source_lang, dest=target_lang)
                    row[target_col - 1].value = translation.text
                except Exception as e:
                    print(f"Error translating '{source_text}': {e}")

    # Save the updated file with a language suffix
    base_name, ext = os.path.splitext(file_path)
    output_file = f"{base_name}_{target_lang}{ext}"
    workbook.save(output_file)
    print(f"Translations saved in {output_file}")

# Parameters (Modify these as needed)
file_path = "translations.xlsx"  # Path to your Excel file
columns_to_translate = [1, 3]  # List of columns to translate (1-based indexing)
source_lang = "en"  # Source language (English)
target_lang = "es"  # Target language (e.g., Spanish)
replace = False  # Set True to replace the source columns

# Translate the specified columns
translate_excel(file_path, columns_to_translate, source_lang, target_lang, replace=replace)
